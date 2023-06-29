import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.styles import numbers
import logging

logger = logging.getLogger(__name__)


class TagNotFound(Exception):
    """
    Exception raised for excel template not containing tags.

    Attributes:
        tag - the tag contained in the Excel template.
        message - the message printed when exception is raised.
        sheet - the Excel sheet missing the tag.
    """

    def __init__(self, tag, sheet): # add sheet
        self.tag = tag
        self.sheet = sheet
        self.message = f"{tag} not found in sheet: {sheet}. Please amend the Excel template."
        super().__init__(self.message)


def find_starting_cell(wb, sheet, tag):
    """Return the row and cell of the tag.
    """

    ws = wb[sheet]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == tag:
                return coordinate_to_tuple(cell.coordinate)
    else:
        raise TagNotFound(tag, sheet)

def prepare_lod_data(csv_path, section):
    logger.info(f"Preparing data for Length of Detention {section}")
    df = pd.read_csv(csv_path)
    df = df[df["section"] == section]
    df = df[["MHA_MOST_SEVERE_CATEGORY", "MHA_Most_severe_Section", "DEMOGRAPHIC", "DEMOGRAPHIC_CATEGORY", "COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]

    return df

def prepare_episode_data(csv_path):
    logger.info("Preparing data for episode data as part of a continous MHA spell")
    df = pd.read_csv(csv_path)
    group_df = df.groupby("section").sum().reset_index()
    group_df["EPISODE_COUNT"] = "Total"
    group_df = group_df[["section", "EPISODE_COUNT", "COUNT"]]
    new_df = pd.concat([df, group_df])
    pivot_df = pd.pivot_table(new_df, values="COUNT", index="EPISODE_COUNT", columns="section", aggfunc="sum").reset_index().set_index("EPISODE_COUNT")
    mha_total = pivot_df.loc["Total", "MHA only"]
    mha_cto_total = pivot_df.loc["Total", "MHA and CTO"]
    pivot_df["MHA only %"] = round((pivot_df["MHA only"] / mha_total) * 100, 1)
    pivot_df["MHA and CTO %"] = round((pivot_df["MHA and CTO"] / mha_cto_total) * 100, 1)
    pivot_df = pivot_df[["MHA only", "MHA and CTO", "MHA only %", "MHA and CTO %"]]
    rows = ["Total", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10+"]
    pivot_df = pivot_df.loc[rows]

    return pivot_df

def prepare_table1a_data(csv_path):
    logger.info(f"Preparing data for MHA Main Table 1a")
    # This list sets out the order in which we want the values to appear
    measures = ["All detentions",
            "Detentions on admission to hospital",
            "Detentions following admission to hospital",
            "Detentions following use of Place of Safety Order",
            "Detentions following revocation of CTO"
            ]   
    df = pd.read_csv(csv_path)
    mask = (df["Measure"].isin(measures)) & (df["Demographic"] == "All") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df.astype({"Total_Count": "int64"}) ##needs to be integer when inserting values into excel

    return df

def prepare_age_gender_rates_data(df, measure):
    breakdowns = ["All", "Age", "Gender"]
    mask = (df["breakdown"].isin(breakdowns)) & (df["count_of"] == measure)
    df = df[mask]    

    return df

def prepare_eth_rates_data(df, measure):
    breakdowns = ["All", "Ethnicity"]
    mask = (df["breakdown"].isin(breakdowns)) & (df["count_of"] == measure)
    df = df[mask]    

    return df

def prepare_crosstab_data(csv_path, all_ages):
    df = pd.read_csv(csv_path)
    df = df[df["Age"].isin(all_ages)]

    return df

def prepare_stp_rates_data(df, measure):
    breakdowns = ["All", "STP"]
    mask = (df["breakdown"].isin(breakdowns)) & (df["count_of"] == measure)
    df = df[mask]    

    return df

def prepare_ccg_rates_data(df, measure):
    breakdowns = ["All", "CCG"]
    mask = (df["breakdown"].isin(breakdowns)) & (df["count_of"] == measure)
    df = df[mask]    

    return df

def prepare_imd_rates_data(df, measure):
    logger.info(f"Preparing data for MHA Main Table 1g")
    breakdowns = ["All", "IMD"]
    mask = (df["breakdown"].isin(breakdowns)) & (df["count_of"] == measure)
    df = df[mask]    

    return df

def prepare_table2a_data(csv_path):
    logger.info(f"Preparing data for MHA Main Table 2a")
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "Short term orders") & (df["Demographic"] == "All") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df.astype({"Total_Count": "int64"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table3a_data(csv_path):
    logger.info(f"Preparing data for MHA Main Table 3a")
    # This list sets out the order in which we want the values to appear
    measures = ["Uses of CTOs",
            "CTO recalls to hospital",
            "Revocations of CTO",
            "Discharges from CTO"
            ]   
    df = pd.read_csv(csv_path)
    mask = (df["Measure"].isin(measures)) & (df["Demographic"] == "All") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df.astype({"Total_Count": "int64"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table4_data(csv_path):
    logger.info(f"Preparing data for MHA Main Table 4")
    # This list sets out the order in which we want the values to appear
    measures = ["Uses of section 2",
            "Uses of section 3"
            ]   
    df = pd.read_csv(csv_path)
    mask = (df["Measure"].isin(measures)) & (df["Demographic"] == "All") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df.astype({"Total_Count": "int64"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table5_data(csv_path):
    logger.info(f"Preparing provider data for MHA Main Table 5")
    measures = ["People detained in hospital on 31st March",
    "People subject to Community Treatment Orders (CTOs) on 31st March", 
    "People subject to the Act on 31st March"
    ]
    df = pd.read_csv(csv_path)
    df = df[df["Measure"].isin(measures)]

    return df

def prepare_table5_provider_data(xlsx_path):
    logger.info(f"Preparing provider data for MHA Main Table 5")
    df = pd.read_excel(xlsx_path, "Table 5")
    df = df.iloc[:, 1].to_frame("OrgID")
    first_idx = df.first_valid_index()
    last_idx = df.last_valid_index()
    df = df.loc[first_idx:last_idx] ##takes off blank values at both ends of dataframe
    df = df.fillna(0) #replace NaN with 0
    
    return df

def prepare_table6_data(csv_path):
    logger.info(f"Preparing data for MHA Main Table 6")    
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "People subject to repeat detentions") & (df["OrganisationBreakdown"] == "All submissions")
    df = df[mask]

    return df

def prepare_pop_data(csv_path):
    logger.info(f"Preparing data for Population data")
    df = pd.read_csv(csv_path)
    df = df[df["count_of"] == "All detentions"]
    df = df[["breakdown","primary_level", "population"]]
    df.columns = ["Breakdown","DemographicBreakdown", "population"]
    pop_df = df.astype({'Breakdown': 'str', 'DemographicBreakdown': 'str', 'population': 'int64'})

    return pop_df

def prepare_demog_repeat_detentions(df, demog):
    df = df[df["Demographic"] == demog]
    df = pd.pivot_table(df, values="Total_Count", index="DemographicBreakdown", columns="measureSubcategory", aggfunc="sum").reset_index()
    df = df.astype({
    '1': 'int64', 
    '2': 'int64', 
    '3': 'int64', 
    '4': 'int64', 
    '5': 'int64', 
    '6': 'int64', 
    '7 and over': 'int64'})
    df["total"] = df.sum(axis=1)
    onedet = df.loc[:, "1"]
    df["morethan1detperc"] = round(((df["total"] - onedet) / df["total"]) * 100, 1)

    return df

def prepare_all_pop(df):
    mask = (df["Breakdown"] == "All") & (df["DemographicBreakdown"] == "All")
    df = df[mask]
    
    return df

def prepare_demog_pop(df, demog):
    df = df[df["Breakdown"] == demog]
    
    return df

def prepare_table6_demog(df, pop_df, demog):
    df = prepare_demog_repeat_detentions(df, demog)
    pop_df = prepare_demog_pop(pop_df, demog)
    df = df.merge(pop_df, on="DemographicBreakdown", how="left")
    df = df[["DemographicBreakdown", "1", "2", "3", "4", "5", "6", "7 and over", "morethan1detperc", "population"]]
    df = df.set_index("DemographicBreakdown")
    
    return df

def prepare_table_1a_1(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a All detentions")
    mask = (df["Measure"] == "All detentions") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_2(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Total Detentions on admission to hospital")
    mask = (df["Measure"] == "Detentions on admission to hospital") & (df["measureSubcategory"] == "All") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_3(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Part2 Detentions")
    mask = (df["Measure"] == "Detentions on admission to hospital") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All detentions under Part 2",
        "Section 2",
        "Section 3"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_4(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Part3 Detentions")
    mask = (df["Measure"] == "Detentions on admission to hospital") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All detentions under Part 3",
        "Section 35",
        "Section 36",
        "Section 37 with S41 restrictions",
        "Section 37",
        "Section 45A",
        "Section 47 with S49 restrictions",
        "Section 47",
        "Section 48 with S49 restrictions",
        "Section 48",
        "Detentions under other sections (38,44 and 46)"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_5(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Detentions under other acts")
    mask = (df["Measure"] == "Detentions on admission to hospital") & (df["measureSubcategory"] == "Detentions under other acts") & (df["OrganisationBreakdown"] == provider_type) 
    df = df[mask]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_6(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Detentions following admission to hospital")
    mask = (df["Measure"] == "Detentions following admission to hospital") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All",
        "Informal to section 2",
        "Informal to section 3",
        "Section 5(2) to 2",
        "Section 5(2) to 3",
        "Section 5(4) to 2",
        "Section 5(4) to 3",
        "Section 4 to 2",
        "Section 4 to 3"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_7(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Detentions following use of Place of Safety Order")
    mask = (df["Measure"] == "Detentions following use of Place of Safety Order") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All",
        "Section 135 to 2",
        "Section 135 to 3",
        "Section 136 to 2",
        "Section 136 to 3"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_1a_8(df, provider_type):
    logger.info(f"Preparing data for excel tag table 1a Detentions following revocation of CTO")
    mask = (df["Measure"] == "Detentions following revocation of CTO") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    df = df[["Total_Count"]]

    return df

def prepare_table_age_gender_1(df):
    logger.info(f"Preparing data for excel tag All detentions, population and crude rate")    
    df = df[df["breakdown"] == "All"]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_age_gender_2(df):
    logger.info(f"Preparing data for excel tag All ages detentions, population and crude rate")
    mask = (df["breakdown"] == "Age") & (df["primary_level_desc"] == "All")     
    df = df[mask]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_age_all(df, age_list):
    logger.info(f"Preparing data for excel tag {age_list[0]} ages detentions, population and crude rate")
    df = df[(df["primary_level_desc"].isin(age_list))]
    df = df.reset_index().set_index('primary_level_desc')
    df = df.loc[age_list]
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_gender_all(df, gen_list):
    logger.info(f"Preparing data for excel tag All genders detentions, population and crude rate")   
    mask = (df["breakdown"] == "Gender") & (df["primary_level_desc"].isin(gen_list))     
    df = df[mask]
    df = df.reset_index().set_index('primary_level_desc')
    df = df.loc[gen_list]
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_eth_1(df):
    logger.info(f"Preparing data for excel tag All detentions, population, crude rate, standardised rate and confidence interval")    
    df = df[df["breakdown"] == "All"]    
    df = df[["count", "population", "CR", "SR", "CI"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float", "SR": "float", "CI": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_eth_2(df):
    logger.info(f"Preparing data for excel tag All Ethnicities detentions, population, crude rate, standardised rate and confidence interval")
    mask = (df["breakdown"] == "Ethnicity") & (df["primary_level_desc"] == "All")     
    df = df[mask]    
    df = df[["count", "population", "CR", "SR", "CI"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float", "SR": "float", "CI": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_eth_lower(df, lower_eth_list):
    logger.info(f"Preparing data for excel tag {lower_eth_list[0]} ethnicities detentions, population, crude rate, standardised rate and confidence interval") 
    mask = (df["breakdown"] == "Ethnicity") & (df["primary_level_desc"].isin(lower_eth_list))     
    df = df[mask]
    df = df.reset_index().set_index('primary_level_desc')
    df = df.loc[lower_eth_list]
    df = df[["count", "population", "CR", "SR", "CI"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float", "SR": "float", "CI": "float"}) ##needs to be integer when inserting values into excel

    return df

def table_1e_crosstab_total(df, col_list, col_name, count_of):
    df = df.groupby([col_name]).sum().reset_index()
    df["CrudeRate"] = round((df["Count"] / df["Population"]) * 100000, 1)
    df = df.replace(np.inf, 0)
    p_df = pd.pivot_table(df, values=count_of, columns=col_name, aggfunc="sum")
    p_df = p_df[col_list]
    
    return p_df

def table_1e_crosstab(df, col_list, col_name, index_list, index_name, count_of): 
    
    df = df[df[index_name].isin(index_list)]
    df = df.groupby([index_name, col_name]).sum().reset_index()
    df["CrudeRate"] = round((df["Count"] / df["Population"]) * 100000, 1)
    df = df.replace(np.inf, 0)
    p_df = pd.pivot_table(df, values=count_of, index=index_name, columns=col_name, aggfunc="sum")
    p_df = p_df.loc[index_list]
    p_df = p_df[col_list]
    
    return p_df

def table_1e_crosstab_same(df, col_list, same, count_of):    
    df = df.groupby(same).sum().reset_index()
    df["CrudeRate"] = round((df["Count"] / df["Population"]) * 100000, 1)
    df = df.merge(df, on=same)
    c_df = pd.crosstab(df[same], df[same], values=df[f"{count_of}_x"], aggfunc="sum")
    c_df = c_df.loc[col_list]
    c_df = c_df[col_list]
    c_df = c_df.replace(np.nan, 0)
    c_df = c_df.replace(np.inf, 0)
    
    return c_df

def prepare_table_stp_1(df):
    logger.info(f"Preparing data for excel tag All detentions, population, crude rate")    
    df = df[df["breakdown"] == "All"]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_stp_2(df):
    logger.info(f"Preparing data for excel tag All STP detentions, population, crude rate")    
    mask = (df["breakdown"] == "STP") & (df["primary_level_desc"] == "All prov") & (df["count_of"] == "All detentions")    
    df = df[mask]
    df = df.astype({"count": "float"})    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_stp_3(df):
    logger.info(f"Preparing data for excel tag All detentions, population, crude rate at STP")    
    mask1 = (df["breakdown"] == "STP") & (~df["primary_level_desc"].isin(["All prov", "UNKNOWN"]))
    df1 = df[mask1]
    df1 = df1[["primary_level", "primary_level_desc", "count", "population", "CR"]]
    df1 = df1.sort_values("primary_level_desc")
    mask2 = (df["breakdown"] == "STP") & (df["primary_level_desc"] == "UNKNOWN")
    df2 = df[mask2]
    df2 = df2[["primary_level", "primary_level_desc", "count", "population", "CR"]]
    df = pd.concat([df1, df2])
    df = df.astype({"primary_level": "object", "primary_level_desc": "object", "count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_ccg_1(df):
    logger.info(f"Preparing data for excel tag All detentions, population, crude rate")    
    df = df[df["breakdown"] == "All"]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_ccg_2(df):
    logger.info(f"Preparing data for excel tag All STP detentions, population, crude rate")    
    mask = (df["breakdown"] == "CCG") & (df["primary_level_desc"] == "All prov")     
    df = df[mask]    
    df = df.astype({"count": "float"})
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_ccg_3(df):
    logger.info(f"Preparing data for excel tag All detentions, population, crude rate at STP")    
    mask1 = (df["breakdown"] == "CCG") & (~df["primary_level_desc"].isin(["All prov", "UNKNOWN"]))
    df1 = df[mask1]
    df1 = df1[["primary_level", "primary_level_desc", "count", "population", "CR"]]
    df1 = df1.sort_values("primary_level_desc")
    mask2 = (df["breakdown"] == "CCG") & (df["primary_level_desc"] == "UNKNOWN")
    df2 = df[mask2]
    df2 = df2[["primary_level", "primary_level_desc", "count", "population", "CR"]]
    df = pd.concat([df1, df2])
    df = df.astype({"primary_level": "object", "primary_level_desc": "object", "count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_1g_1(df):
    logger.info(f"Preparing data for excel tag table 1g All detentions, population and crude rate")    
    df = df[df["breakdown"] == "All"]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_1g_2(df):
    logger.info(f"Preparing data for excel tag table 1g All IMD Deciles detentions, population and crude rate")
    mask = (df["breakdown"] == "IMD") & (df["primary_level_desc"] == "All")     
    df = df[mask]    
    df = df[["count", "population", "CR"]]
    df = df.astype({"count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_1g_imd_upper(df, imd_list):
    logger.info(f"Preparing data for excel tag table 1g IMD Deciles detentions, population and crude rate") 
    mask = (df["breakdown"] == "IMD") & (df["primary_level_desc"].isin(imd_list))     
    df = df[mask]
    df = df.reset_index().set_index('primary_level_desc')
    df = df.loc[imd_list]
    df = df.reset_index()
    df = df[["primary_level_desc", "count", "population", "CR"]]
    df = df.astype({"primary_level_desc": "object", "count": "int64", "population": "int64", "CR": "float"}) ##needs to be integer when inserting values into excel

    return df

def prepare_table_1h_eth_lower(df, lower_eth_list):
    logger.info(f"Preparing data for excel tag table 1h {lower_eth_list[0]} ethnicity IMD")
    df = df[df["DemographicBreakdown"].isin(lower_eth_list)]
    p_df = pd.pivot_table(df, values="COUNT", index="DemographicBreakdown", columns="IMD", aggfunc="sum")
    p_df = p_df.loc[lower_eth_list] #count df
    pp_df = round(p_df.div(p_df.sum(axis=1), axis=0), 3) #percentage df

    return p_df, pp_df

def prepare_stp_demog(csv_path, breakdown, demog_list):
    df = pd.read_csv(csv_path)
    mask = (df["breakdown"] == breakdown) & (df["count_of"] == "All detentions") & (df["primary_level_desc"].isin(demog_list))
    df = df[mask]
    df = df.groupby("primary_level_desc").agg({"count": "sum", "population": "sum", "CR": "sum"})
    df = df.transpose()
    df = df[demog_list]

    return df

def prepare_stp_demog_lower(csv_path, breakdown, demog_list):
    df = pd.read_csv(csv_path)
    mask = (df["breakdown"] == breakdown) & (df["count_of"] == "All detentions")
    df = df[mask]
    df = df[["primary_level", "primary_level_desc", "secondary_level_desc", "count", "population", "CR"]]
    df.columns = ["primary_level", "primary_level_desc", "secondary_level_desc", "Number of detentions", "Base population", "Crude rate per 100,000"]
    df = df.groupby(["primary_level", "primary_level_desc", "secondary_level_desc"]).agg({"Number of detentions": "sum", "Base population": "sum", "Crude rate per 100,000": "sum"})
    df = df.transpose()
    new_df = pd.DataFrame()
    start_col = ["stp_code", "stp_name", "index"]
    final_cols = start_col + demog_list
    stp_names = [stp_name for stp_name in sorted(df.columns.get_level_values(1).unique()) if stp_name != "UNKNOWN"]
    #all stps except UNKNOWN
    for stp_name in stp_names:
        l_df = df.iloc[:, df.columns.get_level_values(1)==stp_name] 
        stp_code_col = l_df.columns[0][0]
        stp_name_col = l_df.columns[0][1]       
        l_df.columns = l_df.columns.droplevel("primary_level") #drop stp code from column index
        l_df.columns = l_df.columns.droplevel("primary_level_desc") #drop stp name from column index 
        l_df["stp_code"] = stp_code_col
        l_df["stp_name"] = stp_name_col
        l_df = l_df.reset_index(level=0)
        l_df = l_df[final_cols]
        new_df = pd.concat([new_df, l_df])
    #unknown stp
    l_df = df.iloc[:, df.columns.get_level_values(0)=="UNKNOWN"] 
    stp_code_col = l_df.columns[0][0]
    stp_name_col = l_df.columns[0][1]       
    l_df.columns = l_df.columns.droplevel("primary_level") #drop stp code from column index
    l_df.columns = l_df.columns.droplevel("primary_level_desc") #drop stp name from column index 
    l_df["stp_code"] = stp_code_col
    l_df["stp_name"] = stp_name_col
    l_df = l_df.reset_index(level=0)
    l_df = l_df[final_cols]
    new_df = pd.concat([new_df, l_df]) 

    return new_df

def combine_stp_demog_lower(csv_path, breakdown1, breakdown2, demog_list1, demog_list2):
    df1 = prepare_stp_demog_lower(csv_path, breakdown1, demog_list1)
    df2 = prepare_stp_demog_lower(csv_path, breakdown2, demog_list2)
    df_comb = df1.merge(df2, on=["stp_code", "stp_name", "index"], how="inner")

    return df_comb

def prepare_table_2a_1(df, provider_type):
    logger.info(f"Preparing data for excel tag table 2a All Short term orders")
    mask = (df["measureSubcategory"] == "All") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_2a_2(df, provider_type):
    logger.info(f"Preparing data for excel tag table 2a All place of safety orders")
    rows = ["All place of safety orders",
        "Section 135",
        "Section 136"
    ]
    mask = (df["OrganisationBreakdown"] == provider_type) & (df["measureSubcategory"].isin(rows))
    df = df[mask]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_2a_3(df, provider_type):
    logger.info(f"Preparing data for excel tag table 2a Uses of section 4")
    mask = (df["measureSubcategory"] == "Section 4") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_2a_4(df, provider_type):
    logger.info(f"Preparing data for excel tag table 2a All uses of section 5")
    rows = ["All uses of section 5",
        "Section 5(2)",
        "Section 5(4)"
    ]
    mask = (df["OrganisationBreakdown"] == provider_type) & (df["measureSubcategory"].isin(rows))
    df = df[mask]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_3a_1(df, provider_type):
    logger.info(f"Preparing data for excel tag table 3a Uses of Community Treatment Orders")
    mask = (df["Measure"] == "Uses of CTOs") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All",
        "Section 3 to CTO",
        "Section 37 to CTO",
        "Section 47 to CTO",
        "Section 48 to CTO",
        "Informal to CTO",
        "Other sections to CTO"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_3a_2(df, provider_type):
    logger.info(f"Preparing data for excel tag table 3a CTO recalls to hospital")
    mask = (df["Measure"] == "CTO recalls to hospital") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_3a_3(df, provider_type):
    logger.info(f"Preparing data for excel tag table 3a Revocations of CTO")
    mask = (df["Measure"] == "Revocations of CTO") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_3a_4(df, provider_type):
    logger.info(f"Preparing data for excel tag table 3a Discharges from CTO")
    mask = (df["Measure"] == "Discharges from CTO") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]    
    df = df[["Total_Count"]]

    return df

def prepare_table_4_1(df, provider_type):
    logger.info(f"Preparing data for excel tag table 4 Uses of Section 2")
    mask = (df["Measure"] == "Uses of section 2") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All",
        "Section 2 on admission",
        "Section 2 following admission (informal to s 2)",
        "Section 5(2) to 2",
        "Section 5(4) to 2",
        "Section 135 to 2",
        "Section 136 to 2",
        "Section 4 to 2"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table_4_2(df, provider_type):
    logger.info(f"Preparing data for excel tag table 4 Uses of Section 2")
    mask = (df["Measure"] == "Uses of section 3") & (df["OrganisationBreakdown"] == provider_type)
    df = df[mask]
    rows = ["All",
        "Section 3 on admission",
        "Section 3 following admission (informal to s3)",
        "Section 5(2) to 3",
        "Section 5(4) to 3",
        "Section 135 to 3",
        "Section 136 to 3",
        "Section 4 to 3",
        "Section 2 to 3",
        "Changes from other section to 3"
    ]
    df = df[(df["measureSubcategory"].isin(rows))]
    df = df.reset_index().set_index('measureSubcategory')
    df = df.loc[rows]
    df = df[["Total_Count"]]

    return df

def prepare_table5_all_prov(df, measure):
    logger.info(f"Preparing data for excel tag All submissions {measure}")
    mask = (df["Measure"] == measure) & (df["OrgID"] == "All prov") & (df["OrganisationBreakdown"] == "All submissions")
    df = df[mask]
    df = df[["Total_Count"]]

    return df

def prepare_table5_prov_type(df, measure):
    logger.info(f"Preparing data for excel tag Provider type {measure}")
    mask = (df["Measure"] == measure) & (df["OrgID"] == "All prov") & (df["OrganisationBreakdown"] != "All submissions")
    df = df[mask]
    df = df[["OrganisationBreakdown", "Total_Count"]]
    df = df.set_index("OrganisationBreakdown")
    prov_types = ["NHS TRUST", "INDEPENDENT HEALTH PROVIDER"]
    df = df.loc[prov_types]

    return df

def prepare_table5_providers(df, measure, prov_df):
    logger.info(f"Preparing data for excel tag Provider type {measure}")
    mask = (df["Measure"] == measure) & (df["OrgID"] != "All prov")
    df = df[mask]
    df = df[["OrgID", "Total_Count"]]
    df = prov_df.merge(df, how="left", on="OrgID")
    df = df.fillna("-")
    df = df.set_index("OrgID")

    return df

def prepare_table6_1(df, pop_df):
    logger.info(f"Preparing data for repeat detentions excel tag All demographics")
    df = prepare_demog_repeat_detentions(df, "All")
    pop_df = prepare_all_pop(pop_df)
    df = df.merge(pop_df, on="DemographicBreakdown", how="left")
    df = df[["DemographicBreakdown", "1", "2", "3", "4", "5", "6", "7 and over", "morethan1detperc", "population"]]
    df = df.fillna(0)
    df = df.set_index("DemographicBreakdown")
    
    return df

def prepare_table6_eth_lower(df, pop_df, demog, lower_eth_list):
    logger.info(f"Preparing data for repeat detentions excel tag {lower_eth_list[0]} Ethnicities")
    df = prepare_table6_demog(df, pop_df, demog)
    df = df.loc[lower_eth_list]
    
    return df

def prepare_lod_all_all(df):
    logger.info(f"Preparing data for los excel tag All detentions All demographics")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "All Detentions") & (df["DEMOGRAPHIC"] == "All")
    df = df[mask]
    df = df.groupby(["MHA_MOST_SEVERE_CATEGORY", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df[["COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]
    df = df.astype("float")
    df = df.transpose()
    df.columns = df.columns.droplevel()
    df = pd.DataFrame(df.values.astype("int"))
    df.columns = ["FIG"]
    df = df[["FIG"]]   

    return df

def prepare_lod_all_part2(df):
    logger.info(f"Preparing data for los excel tag Part II Detentions All demographics")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Part II") & (df["DEMOGRAPHIC"] == "All")
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 2",
    "Formally detained under Mental Health Act Section 3",
    "Formally detained under Mental Health Act Section 4",
    "Formally detained under Mental Health Act Section 5(2)",
    "Formally detained under Mental Health Act Section 5(4)",
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df[["COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]
    df = df.astype("float")
    df = df.transpose()
    df = pd.Series(df.values.astype("int").ravel("F")).to_frame("FIG")
    df = df.reset_index()
    df = df[["FIG"]]

    return df

def prepare_lod_all_court(df):
    logger.info(f"Preparing data for los excel tag Court and prison disposals All demographics")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Court and prison disposals") & (df["DEMOGRAPHIC"] == "All")
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 35",
    "Formally detained under Mental Health Act Section 36",
    "Formally detained under Mental Health Act Section 37 with section 41 restrictions",
    "Formally detained under Mental Health Act Section 37",
    "Formally detained under Mental Health Act Section 38",
    "Formally detained under Mental Health Act Section 45A",
    "Formally detained under Mental Health Act Section 46",
    "Formally detained under Mental Health Act Section 47 with section 49 restrictions",
    "Formally detained under Mental Health Act Section 47",
    "Formally detained under Mental Health Act Section 48 with section 49 restrictions",
    "Formally detained under Mental Health Act Section 48",
    "Subject to guardianship under Mental Health Act Section 7",
    "Subject to guardianship under Mental Health Act Section 37",
    "Formally detained under other acts",
    "Formally detained under Criminal Proceedings (Insanity) Act 1964 as amended by the Criminal Procedures (Insanity and Unfitness to Plead) Act 1991"
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df[["COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]
    df = df.astype("float")
    df = df.transpose()
    df = pd.Series(df.values.astype("int").ravel("F")).to_frame("FIG")
    df = df.reset_index()
    df = df[["FIG"]]

    return df

def prepare_lod_all_pos(df):
    logger.info(f"Preparing data for los excel tag Place of Safety All demographics")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Place of safety") & (df["DEMOGRAPHIC"] == "All")
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 135",
    "Formally detained under Mental Health Act Section 136"
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df[["COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]
    df = df.astype("float")
    df = df.transpose()
    df = pd.Series(df.values.astype("int").ravel("F")).to_frame("FIG")
    df = df.reset_index()
    df = df[["FIG"]]

    return df

def prepare_lod_all_cto(df):
    logger.info(f"Preparing data for los excel tag Community Treatment Orders All demographics")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Civil Community-based Detention") & (df["DEMOGRAPHIC"] == "All")
    df = df[mask]
    sections = ["Community Treatment Order"]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df[["COUNT", "LOWER_QUARTILE", "MEDIAN", "UPPER_QUARTILE"]]
    df = df.astype("float")
    df = df.transpose()
    df = pd.Series(df.values.astype("int").ravel("F")).to_frame("FIG")
    df = df.reset_index()
    df = df[["FIG"]]

    return df

def prepare_lod_demog_all(df, demog, demog_list):
    logger.info(f"Preparing data for los excel All detentions {demog}")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "All Detentions") & (df["DEMOGRAPHIC"] == demog)
    df = df[mask]
    df = df.groupby(["MHA_MOST_SEVERE_CATEGORY", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.astype("float")
    df = df.transpose()
    df.columns = df.columns.droplevel()
    df = df[demog_list]
    df = pd.DataFrame(df.values.astype("int"))

    return df

def prepare_lod_demog_part2(df, demog, demog_list):
    logger.info(f"Preparing data for los excel tag Part II detentions {demog}")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Part II") & (df["DEMOGRAPHIC"] == demog)
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 2",
    "Formally detained under Mental Health Act Section 3",
    "Formally detained under Mental Health Act Section 4",
    "Formally detained under Mental Health Act Section 5(2)",
    "Formally detained under Mental Health Act Section 5(4)",
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df.transpose()
    new_df = pd.DataFrame()
    for column in demog_list:
        l_df = df.iloc[:, df.columns.get_level_values(1)==column]
        new_df[column] = pd.Series(l_df.values.astype("int").ravel("F"))

    return new_df

def prepare_lod_demog_court(df, demog, demog_list):
    logger.info(f"Preparing data for los excel tag Court and prison disposals {demog}")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Court and prison disposals") & (df["DEMOGRAPHIC"] == demog)
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 35",
    "Formally detained under Mental Health Act Section 36",
    "Formally detained under Mental Health Act Section 37 with section 41 restrictions",
    "Formally detained under Mental Health Act Section 37",
    "Formally detained under Mental Health Act Section 38",
    "Formally detained under Mental Health Act Section 45A",
    "Formally detained under Mental Health Act Section 46",
    "Formally detained under Mental Health Act Section 47 with section 49 restrictions",
    "Formally detained under Mental Health Act Section 47",
    "Formally detained under Mental Health Act Section 48 with section 49 restrictions",
    "Formally detained under Mental Health Act Section 48",
    "Subject to guardianship under Mental Health Act Section 7",
    "Subject to guardianship under Mental Health Act Section 37",
    "Formally detained under other acts",
    "Formally detained under Criminal Proceedings (Insanity) Act 1964 as amended by the Criminal Procedures (Insanity and Unfitness to Plead) Act 1991"
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df.transpose()
    new_df = pd.DataFrame()
    for column in demog_list:
        l_df = df.iloc[:, df.columns.get_level_values(1)==column]
        new_df[column] = pd.Series(l_df.values.astype("int").ravel("F"))

    return new_df

def prepare_lod_demog_pos(df, demog, demog_list):
    logger.info(f"Preparing data for los excel tag Place of safety {demog}")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Place of safety") & (df["DEMOGRAPHIC"] == demog)
    df = df[mask]
    sections = ["All",
    "Formally detained under Mental Health Act Section 135",
    "Formally detained under Mental Health Act Section 136"
    ]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.loc[sections]
    df = df.transpose()
    new_df = pd.DataFrame()
    for column in demog_list:
        l_df = df.iloc[:, df.columns.get_level_values(1)==column]
        new_df[column] = pd.Series(l_df.values.astype("int").ravel("F"))

    return new_df

def prepare_lod_demog_cto(df, demog, demog_list):
    logger.info(f"Preparing data for los excel tag Community Treatment Orders {demog}")
    mask = (df["MHA_MOST_SEVERE_CATEGORY"] == "Civil Community-based Detention") & (df["DEMOGRAPHIC"] == demog)
    df = df[mask]
    sections = ["Community Treatment Order"]
    df = df[df["MHA_Most_severe_Section"].isin(sections)]
    df = df.groupby(["MHA_Most_severe_Section", "DEMOGRAPHIC_CATEGORY"]).sum()
    df = df.transpose()
    new_df = pd.DataFrame()
    for column in demog_list:
        l_df = df.iloc[:, df.columns.get_level_values(1)==column]
        new_df[column] = pd.Series(l_df.values.astype("int").ravel("F"))

    return new_df

def prepare_dq_prov_subm_det_data(csv_path, value):
    logger.info(f"Preparing data for DQ {value} by provider type")    
    df = pd.read_csv(csv_path)
    p_df = pd.pivot_table(df, values=value, index="year", columns="org_type", aggfunc="sum")
    p_df = p_df[["MH & LD", "ACUTE", "INDEPENDENT"]]

    return p_df

def prepare_dq_prov_comp_data(csv_path, value):
    logger.info(f"Preparing data for {value} Provider completeness over 12 months")    
    df = pd.read_csv(csv_path)
    df = df[df["OrgType"] == value]
    p_df = pd.pivot_table(df, values="OrgIDProvider", columns="Submissions", aggfunc="count")
    p_df = p_df[["Less than 12 months", "All 12 months"]]

    return p_df

def prepare_mhs08_data(csv_path):
    logger.info(f"Preparing data for MHS08 values over 12 months")    
    df = pd.read_csv(csv_path)
    df = df[df["BREAKDOWN"] == "England"]
    df = df[["REPORTING_PERIOD_END", "MEASURE_VALUE"]]
    df = df.astype({"REPORTING_PERIOD_END": "datetime64[ns]", "MEASURE_VALUE": "int32"})
    df["MONTH_YEAR"] = df["REPORTING_PERIOD_END"].dt.strftime("%b-%y")
    df = df.groupby(["MONTH_YEAR", "REPORTING_PERIOD_END"]).sum().reset_index()
    df = df.sort_values("REPORTING_PERIOD_END")
    df = df[["MONTH_YEAR", "MEASURE_VALUE"]]

    return df

def prepare_det_type_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for detention type CMS chart")
    df = pd.read_csv(csv_path)
    measures = ["Detentions on admission to hospital",
            "Detentions following admission to hospital",
            "Detentions following use of Place of Safety Order",
            "Detentions following revocation of CTO"]
    mask = (df["Measure"].isin(measures)) & (df["measureSubcategory"] == "All") & (df["OrganisationBreakdown"] == "All submissions")
    df = df[mask]
    df = df[["Measure", "Total_Count"]].set_index("Measure")
    df.columns = ["Count"]
    df = df.loc[measures]
    df.to_excel(output_path)

def prepare_det_type_perc_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for detention type percentage CMS chart")
    df = pd.read_csv(csv_path)
    measures = ["Detentions on admission to hospital",
            "Detentions following admission to hospital",
            "Detentions following use of Place of Safety Order",
            "Detentions following revocation of CTO"]
    mask = (df["Measure"].isin(measures)) & (df["measureSubcategory"] == "All") & (df["OrganisationBreakdown"] != "All submissions") & (df["OrgID"] == "All prov")
    df = df[mask]
    df["DetentionType"] = np.where(df["Measure"] == "Detentions on admission to hospital", "On admission",
                        np.where(df["Measure"] == "Detentions following admission to hospital", "Following admission", 
                                "Other")) #else
    df["ProviderType"] = np.where(df["OrganisationBreakdown"] == "NHS TRUST", "NHS Providers", "Independent Providers") #else
    df = df[["DetentionType", "ProviderType", "Total_Count"]]
    df = df.astype({"DetentionType": "object", "ProviderType": "object", "Total_Count": "int32"})
    df = df.groupby(["ProviderType", "DetentionType"]).sum().reset_index()
    p_df = pd.pivot_table(df, values="Total_Count", index="ProviderType", columns="DetentionType", aggfunc="sum")
    pp_df = round(p_df.div(p_df.sum(axis=1), axis=0) * 100, 1) #percentage df
    pp_df = pp_df.loc[["NHS Providers", "Independent Providers"]]
    pp_df = pp_df[["On admission", "Following admission", "Other"]]
    pp_df.to_excel(output_path)

def prepare_det_eth_count_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for number of detentions by ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["White", "Mixed", "Asian or Asian British", "Black or Black British", "Other Ethnic Groups"]
    mask = (df["Measure"] == "All detentions") & (df["DemographicBreakdown"].isin(eth)) & (df["Demographic"] == "Ethnicity") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df[["DemographicBreakdown", "Total_Count"]]
    df = df.astype({"DemographicBreakdown": "object", "Total_Count": "int32"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[eth].reset_index()
    df.columns = ["Ethnicity", "Count"]
    df.to_excel(output_path, index=False)

def prepare_det_eth_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for standardised rate of detentions by ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["White", "Mixed", "Asian or Asian British", "Black or Black British", "Other Ethnic Groups"]
    mask = (df["Measure"] == "All detentions") & (df["DemographicBreakdown"].isin(eth)) & (df["Demographic"] == "Ethnicity")
    df = df[mask]
    df = df[["DemographicBreakdown", "StdRate", "Confidence_interval"]]
    df = df.astype({"DemographicBreakdown": "object", "StdRate": "float", "Confidence_interval": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[eth].reset_index()
    df.columns = ["Ethnicity", "SR", "CI"]
    df.to_excel(output_path, index=False)

def prepare_det_gender_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for rate of detentions by gender CMS chart")
    df = pd.read_csv(csv_path)
    gen = ["1", "2"]
    mask = (df["Measure"] == "All detentions") & (df["DemographicBreakdown"].isin(gen)) & (df["Demographic"] == "Gender")
    df = df[mask]
    df = df[["DemographicBreakdown", "CrudeRate"]]
    df = df.astype({"DemographicBreakdown": "object", "CrudeRate": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[gen].reset_index()
    df.replace("1", "Male", inplace=True)
    df.replace("2", "Female", inplace=True)
    df.columns = ["Gender", "Crude Rate"]
    df.to_excel(output_path, index=False)

def prepare_det_age_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for rate of detentions by age group CMS chart")
    df = pd.read_csv(csv_path)
    age = ["15 and under", "16 to 17", "18 to 34", "35 to 49", "50 to 64", "65 and over"]
    mask = (df["Measure"] == "All detentions") & (df["DemographicBreakdown"].isin(age)) & (df["Demographic"] == "Age")
    df = df[mask]
    df = df[["DemographicBreakdown", "CrudeRate"]]
    df = df.astype({"DemographicBreakdown": "object", "CrudeRate": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[age].reset_index()
    df.columns = ["Age group", "Crude Rate"]
    df.to_excel(output_path, index=False)

def prepare_det_leth_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for standardised rate of detentions by lower ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["British", "Irish", "Any Other White Background",
        "White and Black Caribbean", "White and Black African", "White and Asian", "Any Other Mixed Background", 
        "Indian", "Pakistani", "Bangladeshi", "Any Other Asian Background", 
        "African", "Caribbean", "Any Other Black Background", 
        "Chinese", "Any other ethnic group"]
    mask = (df["Measure"] == "All detentions") & (df["DemographicBreakdown"].isin(eth)) & (df["Demographic"] == "Ethnicity")
    df = df[mask]
    df = df[["DemographicBreakdown", "StdRate", "Confidence_interval"]]
    df = df.astype({"DemographicBreakdown": "object", "StdRate": "float", "Confidence_interval": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[eth].reset_index()
    df.columns = ["Ethnicity", "SR", "CI"]
    df.to_excel(output_path, index=False)

def prepare_det_imd_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for rate of detentions by imd CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "All detentions") & (df["Demographic"] == "IMD") & (df["DemographicBreakdown"] != "Not stated/Not known/Invalid") & (df["OrgID"] == "All prov")
    df = df[mask]
    df = df[["DemographicBreakdown", "CrudeRate"]]
    df = df.astype({"DemographicBreakdown": "object", "CrudeRate": "float"})
    df = df.groupby("DemographicBreakdown").sum().reset_index()
    df = df.sort_values("DemographicBreakdown")
    df.columns = ["IMD", "Crude Rate"]
    df.to_excel(output_path, index=False)

def prepare_repeat_det_cms_charts(csv_path, count_output_path, perc_output_path):
    logger.info(f"Preparing xlsx file for count and percentage of repeat detentions CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "People subject to repeat detentions") & (df["Demographic"] == "All")
    df = df[mask]
    df["RepeatDetentionType"] = np.where(df["measureSubcategory"] == "1", "Once",
                                np.where(df["measureSubcategory"] == "2", "Twice", "3 or more")) #else
    df = df[["RepeatDetentionType", "Total_Count"]]
    df = df.astype({"RepeatDetentionType": "object", "Total_Count": "int32"})
    df = df.groupby(["RepeatDetentionType"]).sum()
    order = ["Once", "Twice", "3 or more"]
    df = df.loc[order].reset_index()
    df.columns = ["RepeatDetentionType", "Count"]
    #count output
    df.to_excel(count_output_path, index=False)
    #percentage output
    df["Percentage"] = round((df["Count"] / df["Count"].sum()) * 100, 1)
    p_df = df[["RepeatDetentionType", "Percentage"]]
    p_df.to_excel(perc_output_path, index=False) 

def prepare_repeat_det_demog_cms_charts(csv_path, output_path):
    logger.info(f"Preparing xlsx file for count and percentage of repeat detentions Excel chart")
    df = pd.read_csv(csv_path)
    demogs = ["1", "2", "3", "4", "9",
            "15 and under", "16 to 17", "18 to 34", "35 to 49", "50 to 64", "65 and over",
            "White", "Mixed", "Asian or Asian British", "Black or Black British", "Other Ethnic Groups"]
    mask = (df["Measure"] == "People subject to repeat detentions") & (df["DemographicBreakdown"].isin(demogs))
    df = df[mask]
    df["RepeatDetentionType"] = np.where(df["measureSubcategory"] == "1", "Once", "More than once")
    df = df[["DemographicBreakdown", "RepeatDetentionType", "Total_Count"]]
    df = df.astype({"DemographicBreakdown": "object", "RepeatDetentionType": "object", "Total_Count": "int32"})
    p_df = pd.pivot_table(df, values="Total_Count", index="DemographicBreakdown", columns="RepeatDetentionType", aggfunc="sum")
    pp_df = round(p_df.div(p_df.sum(axis=1), axis=0) * 100, 1) #percentage df
    pp_df = pp_df.loc[demogs]
    pp_df = pp_df.reset_index()
    pp_df = pp_df[["DemographicBreakdown", "More than once"]]
    pp_df.replace("1", "Male", inplace=True)
    pp_df.replace("2", "Female", inplace=True)
    pp_df.replace("3", "Non-binary", inplace=True)
    pp_df.replace("4", "Other (not listed)", inplace=True)
    pp_df.replace("9", "Indeterminate", inplace=True)
    pp_df.to_excel(output_path, index=False)

def prepare_median_lod_gender_cms_charts(csv_path, output_path):
    logger.info(f"Preparing xlsx file for median length of detention by gender CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["DEMOGRAPHIC"] == "Gender") & (df["MHA_MOST_SEVERE_CATEGORY"] == "All Detentions")
    df = df[mask]
    p_df = pd.pivot_table(df, values="MEDIAN", index="DEMOGRAPHIC_CATEGORY", columns="section", aggfunc="mean").reset_index()
    p_df.replace("1", "Male", inplace=True)
    p_df.replace("2", "Female", inplace=True)
    p_df.replace("3", "Non-binary", inplace=True)
    p_df.replace("4", "Other (not listed)", inplace=True)
    p_df.replace("9", "Indeterminate", inplace=True)
    p_df.columns = ["Gender", "Including CTOs", "Excluding CTOs"]
    p_df = p_df[["Gender", "Including CTOs", "Excluding CTOs"]]
    p_df.to_excel(output_path, index=False)

def prepare_median_lod_age_cms_charts(csv_path, output_path):
    logger.info(f"Preparing xlsx file for median length of detention by age group CMS chart")
    df = pd.read_csv(csv_path)
    df.replace("15 AND UNDER", "15 and under", inplace=True)
    age = ["15 and under", "16 to 17", "18 to 34", "35 to 49", "50 to 64", "65 and over"]
    mask = (df["DEMOGRAPHIC"] == "Age") & (df["MHA_MOST_SEVERE_CATEGORY"] == "All Detentions") & (df["DEMOGRAPHIC_CATEGORY"].isin(age))
    df = df[mask]
    p_df = pd.pivot_table(df, values="MEDIAN", index="DEMOGRAPHIC_CATEGORY", columns="section", aggfunc="mean").reset_index()
    p_df.columns = ["Age group", "Including CTOs", "Excluding CTOs"]
    p_df = p_df[["Age group", "Including CTOs", "Excluding CTOs"]]
    p_df.to_excel(output_path, index=False)

def prepare_median_lod_eth_cms_charts(csv_path, output_path):
    logger.info(f"Preparing xlsx file for median length of detention by ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["White British", "White Irish", "Any Other White Background",
        "White and Black Caribbean", "White and Black African", "White and Asian", "Any Other Mixed Background", 
        "Indian", "Pakistani", "Bangladeshi", "Any Other Asian Background", 
        "African", "Caribbean", "Any Other Black Background", 
        "Chinese", "Any other ethnic group",
        "Not Stated", "Not Known", "Unknown"]
    mask = (df["DEMOGRAPHIC"] == "Lower Level Ethnicity") & (df["MHA_MOST_SEVERE_CATEGORY"] == "All Detentions") & (df["DEMOGRAPHIC_CATEGORY"].isin(eth))
    df = df[mask]
    p_df = pd.pivot_table(df, values="MEDIAN", index="DEMOGRAPHIC_CATEGORY", columns="section", aggfunc="mean")
    p_df = p_df.loc[eth]
    p_df = p_df.reset_index()
    p_df.columns = ["Ethnicity", "Including CTOs", "Excluding CTOs"]
    p_df.to_excel(output_path, index=False)

def prepare_s136_det_age_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for rate of section 136 detentions by age group CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["measureSubcategory"] == "Section 136") & (df["Demographic"] == "Age")
    df = df[mask]
    age = ["15 and under", "16 to 17", "18 to 34", "35 to 49", "50 to 64", "65 and over"]
    df = df[["DemographicBreakdown", "CrudeRate"]]
    df = df.astype({"DemographicBreakdown": "object", "CrudeRate": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[age].reset_index()
    df.columns = ["Age group", "Crude Rate"]
    df.to_excel(output_path, index=False)

def prepare_s136_det_eth_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for standardised rate of section 136 detentions by ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["White", "Mixed", "Asian or Asian British", "Black or Black British", "Other Ethnic Groups"]
    mask = (df["measureSubcategory"] == "Section 136") & (df["DemographicBreakdown"].isin(eth)) & (df["Demographic"] == "Ethnicity")
    df = df[mask]
    df = df[["DemographicBreakdown", "StdRate", "Confidence_interval"]]
    df = df.astype({"DemographicBreakdown": "object", "StdRate": "float", "Confidence_interval": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[eth].reset_index()
    df.columns = ["Ethnicity", "SR", "CI"]
    df.to_excel(output_path, index=False)

def prepare_cto_age_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for rate of section 136 detentions by age group CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "Uses of CTOs") & (df["Demographic"] == "Age")
    df = df[mask]
    age = ["15 and under", "16 to 17", "18 to 34", "35 to 49", "50 to 64", "65 and over"]
    df = df[["DemographicBreakdown", "CrudeRate"]]
    df = df.astype({"DemographicBreakdown": "object", "CrudeRate": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[age].reset_index()
    df.columns = ["Age group", "Crude Rate"]
    df.to_excel(output_path, index=False)

def prepare_cto_eth_rate_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for standardised rate of section 136 detentions by ethnicity CMS chart")
    df = pd.read_csv(csv_path)
    eth = ["White", "Mixed", "Asian or Asian British", "Black or Black British", "Other Ethnic Groups"]
    mask = (df["Measure"] == "Uses of CTOs") & (df["DemographicBreakdown"].isin(eth)) & (df["Demographic"] == "Ethnicity")
    df = df[mask]
    df = df[["DemographicBreakdown", "StdRate", "Confidence_interval"]]
    df = df.astype({"DemographicBreakdown": "object", "StdRate": "float", "Confidence_interval": "float"})
    df = df.groupby("DemographicBreakdown").sum()
    df = df.loc[eth].reset_index()
    df.columns = ["Ethnicity", "SR", "CI"]
    df.to_excel(output_path, index=False)

def prepare_mhs08_perc_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for percentage of people subjected to the act by provider type CMS chart")
    df = pd.read_csv(csv_path)
    mask = (df["Measure"] == "People subject to the Act on 31st March") & (df["OrgID"] == "All prov") & (df["OrganisationBreakdown"] != "All submissions")
    df = df[mask]
    df["Year"] = df["Year"].str.replace("/", "-")
    p_df = pd.pivot_table(df, values="Total_Count", index="Year", columns="OrganisationBreakdown", aggfunc="sum")
    p_df.columns = ["Detained in independent sector", "Detained in NHS"]
    p_df = p_df.astype("int")
    p_df = p_df[["Detained in NHS", "Detained in independent sector"]]
    pp_df = round(p_df.div(p_df.sum(axis=1), axis=0) * 100, 1) #percentage df
    pp_df.to_excel(output_path)

def prepare_mhs08_prov_type_cms_chart(csv_path, output_path):
    logger.info(f"Preparing xlsx file for percentage of people subjected to the act by provider type CMS chart")
    df = pd.read_csv(csv_path)
    df["OrgType"] = np.where(df["OrgType"] == "NHS TRUST", "NHS Provider",
                    np.where(df["OrgType"] == "INDEPENDENT HEALTH PROVIDER", "Independent Providers", "All providers"))
    df = df[df["MHA_Part"] != "Other"]
    p_df = pd.pivot_table(df, values="mhs09", index="OrgType", columns="MHA_Part", aggfunc="sum")
    p_df.columns = ["Detained under Part II", "Detained under Part III"]
    pp_df = round(p_df.div(p_df.sum(axis=1), axis=0) * 100, 1) #percentage df
    pp_df.to_excel(output_path)

def write_tables_to_excel(tables, excel_template, excel_output):
    """Write data to an excel template.

    Example:
    -------
    write_table_5_4_to_excel(table_data, excel_file=EXCEL_TEMPLATE)
    """
    logger.info(f"Writing tables to excel")
    logger.info(f"Using excel template:\n {excel_template}")
    xl_writer = pd.ExcelWriter(excel_output, engine='openpyxl')
    wb = load_workbook(excel_template)
    xl_writer.book = wb
    xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

    for table in tables:
        start_cell = find_starting_cell(wb, table["sheet_name"], table["tag"])
        table["data"].to_excel(
            xl_writer,
            table["sheet_name"],
            index=False,
            header=False,
            startcol=start_cell[1]-1,
            startrow=start_cell[0]-1
        )
    logger.info(f"Saving outputs to:\n {excel_output}")
    xl_writer.save()
